"""Multi-sheet Excel export for a case: summary, service lines, validation, evidence, audit.

Kept separate from the JSON export (api/main.py's export_package) — same underlying
data, different presentation. A single wide spreadsheet row per claim (~76 CMS-1500
fields plus repeating service lines) is unreadable; splitting into purpose-built
sheets keeps each one scannable.
"""

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def _cell(value: Any) -> Any:
    """openpyxl can't write a dict/list directly into a cell."""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _append_package_rows(
    sheets: dict[str, Worksheet],
    *,
    package_id: str | None,
    fields: list,
    reviewed_data: dict,
    failures: list,
    audit_events: list,
    row_keys: list[str],
) -> None:
    """Appends one package's rows into already-headered sheets. When package_id is
    given, every row is prefixed with it so a batch export can tell packages apart
    on shared sheets."""
    prefix = [package_id] if package_id is not None else []

    for f in fields:
        if f.parent_field is not None:
            continue
        value = json.loads(f.value_json) if f.value_json else None
        if isinstance(value, list):
            continue
        final = reviewed_data.get(f.name, value)
        sheets["summary"].append(
            [*prefix, f.name, _cell(value), _cell(final), f.confidence, f.field_status]
        )

    row_fields = [f for f in fields if f.parent_field is not None]
    for f in row_fields:
        value = json.loads(f.value_json) if f.value_json else {}
        row_values = [
            _cell(value.get(key)) if isinstance(value, dict) else None
            for key in row_keys
        ]
        sheets["lines"].append([*prefix, f.name, *row_values, f.confidence])

    for v in failures:
        sheets["validation"].append(
            [
                *prefix,
                v.field,
                v.rule,
                v.severity,
                v.policy_required,
                v.reason,
                v.machine_value,
                v.expected_value,
            ]
        )

    for f in fields:
        ev = json.loads(f.evidence_json) if f.evidence_json else None
        if ev:
            sheets["evidence"].append(
                [*prefix, f.name, ev.get("page"), ev.get("text"), ev.get("block_type")]
            )

    for e in audit_events:
        detail = json.loads(e.detail_json) if e.detail_json else None
        sheets["audit"].append(
            [
                *prefix,
                e.timestamp.isoformat() if e.timestamp else None,
                e.actor,
                e.action,
                _cell(detail),
            ]
        )


def build_case_workbook(
    *,
    fields: list,
    reviewed_data: dict,
    failures: list,
    audit_events: list,
) -> bytes:
    """fields/failures/audit_events are the ORM rows already fetched by the caller
    (ExtractedField, ValidationFailure, AuditLogEntry) — this function only formats
    them, it doesn't touch the database."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Claim Summary"
    _write_header(
        summary, ["Field", "Extracted Value", "Final Value", "Confidence", "Status"]
    )
    lines = wb.create_sheet("Service Lines")
    row_fields = [f for f in fields if f.parent_field is not None]
    row_keys: list[str] = []
    for f in row_fields:
        value = json.loads(f.value_json) if f.value_json else {}
        if isinstance(value, dict):
            for key in value:
                if key not in row_keys:
                    row_keys.append(key)
    _write_header(lines, ["Row", *row_keys, "Confidence"])
    validation = wb.create_sheet("Validation Results")
    _write_header(
        validation,
        [
            "Field",
            "Rule",
            "Severity",
            "Policy Required",
            "Reason",
            "Machine Value",
            "Expected Value",
        ],
    )
    evidence = wb.create_sheet("Source Evidence")
    _write_header(evidence, ["Field", "Page", "Text", "Block Type"])
    audit = wb.create_sheet("Audit Trail")
    _write_header(audit, ["Timestamp", "Actor", "Action", "Detail"])

    _append_package_rows(
        {
            "summary": summary,
            "lines": lines,
            "validation": validation,
            "evidence": evidence,
            "audit": audit,
        },
        package_id=None,
        fields=fields,
        reviewed_data=reviewed_data,
        failures=failures,
        audit_events=audit_events,
        row_keys=row_keys,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_batch_workbook(packages: list[dict]) -> bytes:
    """One workbook covering multiple packages — each sheet gets a leading Package
    column instead of one workbook per package, so it stays usable past a handful
    of packages. Each dict in `packages`: package_id, fields, reviewed_data,
    failures, audit_events (same shapes as build_case_workbook)."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Claim Summary"
    _write_header(
        summary,
        ["Package", "Field", "Extracted Value", "Final Value", "Confidence", "Status"],
    )
    lines = wb.create_sheet("Service Lines")
    row_keys: list[str] = []
    for pkg in packages:
        for f in pkg["fields"]:
            if f.parent_field is None:
                continue
            value = json.loads(f.value_json) if f.value_json else {}
            if isinstance(value, dict):
                for key in value:
                    if key not in row_keys:
                        row_keys.append(key)
    _write_header(lines, ["Package", "Row", *row_keys, "Confidence"])
    validation = wb.create_sheet("Validation Results")
    _write_header(
        validation,
        [
            "Package",
            "Field",
            "Rule",
            "Severity",
            "Policy Required",
            "Reason",
            "Machine Value",
            "Expected Value",
        ],
    )
    evidence = wb.create_sheet("Source Evidence")
    _write_header(evidence, ["Package", "Field", "Page", "Text", "Block Type"])
    audit = wb.create_sheet("Audit Trail")
    _write_header(audit, ["Package", "Timestamp", "Actor", "Action", "Detail"])

    sheets = {
        "summary": summary,
        "lines": lines,
        "validation": validation,
        "evidence": evidence,
        "audit": audit,
    }
    for pkg in packages:
        _append_package_rows(
            sheets,
            package_id=pkg["package_id"],
            fields=pkg["fields"],
            reviewed_data=pkg["reviewed_data"],
            failures=pkg["failures"],
            audit_events=pkg["audit_events"],
            row_keys=row_keys,
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
