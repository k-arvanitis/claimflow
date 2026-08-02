"""The Excel export must split a case into readable sheets, not one wide row —
see README's Known limitations / the product spec's Excel-export requirement."""

import io
from unittest.mock import MagicMock, patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app

_PDF_BYTES = b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj 2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj 3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj\nxref\n0 4\n0000000000 65535 f\n0000000009 00000 n\n0000000058 00000 n\n0000000115 00000 n\ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n190\n%%EOF"  # noqa: E501


def test_excel_export_splits_case_into_named_sheets():
    fake_result = {
        "package_dir": "/tmp/test",
        "domain": "cms1500",
        "documents": [
            {
                "path": "/tmp/test/claim.pdf",
                "doc_type": "cms1500",
                "has_text_layer": True,
                "scan_quality": None,
            }
        ],
        "extraction_data": {
            "patient_name": "DOE JOHN",
            "service_lines": [{"cpt_code": "99214", "charges": 410.0}],
        },
        "extraction_fields": [
            {
                "name": "patient_name",
                "value": "DOE JOHN",
                "confidence": 1.0,
                "grounded": True,
                "valid": True,
                "field_status": "found",
                "evidence": {
                    "page": 1,
                    "text": "Patient: DOE JOHN",
                    "bbox": None,
                    "block_type": "paragraph",
                },
            },
            {
                "name": "service_lines[0]",
                "value": {"cpt_code": "99214", "charges": 410.0},
                "confidence": 0.9,
                "grounded": True,
                "valid": True,
                "field_status": "found",
                "parent_field": "service_lines",
                "evidence": {
                    "page": 1,
                    "text": "99214 $410",
                    "bbox": None,
                    "block_type": "paragraph",
                },
            },
        ],
        "extraction_status": "review",
        "extraction_overall_confidence": 0.95,
        "validation_failures": [
            {
                "field": "patient_dob",
                "rule": "mandatory",
                "reason": "missing",
                "severity": "error",
                "policy_required": False,
            }
        ],
        "policy_answers": [],
        "decision": "needs_review",
        "review_reasons": ["missing field"],
        "error": None,
    }
    mock_graph = MagicMock()
    mock_graph.invoke.return_value = fake_result

    with patch("api.main.build_graph", return_value=mock_graph):
        with TestClient(app) as client:
            upload = client.post(
                "/packages",
                files=[
                    ("files", ("claim.pdf", io.BytesIO(_PDF_BYTES), "application/pdf"))
                ],
            )
            package_id = upload.json()["package_id"]

            resp = client.get(f"/packages/{package_id}/export.xlsx")
            assert resp.status_code == 200
            assert resp.headers["content-type"] == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            wb = load_workbook(io.BytesIO(resp.content))
            assert wb.sheetnames == [
                "Claim Summary",
                "Service Lines",
                "Validation Results",
                "Source Evidence",
                "Audit Trail",
            ]

            summary_rows = list(wb["Claim Summary"].iter_rows(values_only=True))
            assert summary_rows[0] == (
                "Field",
                "Extracted Value",
                "Final Value",
                "Confidence",
                "Status",
            )
            assert any(
                row[0] == "patient_name" and row[1] == "DOE JOHN"
                for row in summary_rows[1:]
            )
            # The repeating service_lines list is not dumped as JSON into the summary sheet.
            assert not any(row[0] == "service_lines" for row in summary_rows[1:])

            line_rows = list(wb["Service Lines"].iter_rows(values_only=True))
            assert line_rows[0][0] == "Row"
            assert "cpt_code" in line_rows[0]
            assert any(row[0] == "service_lines[0]" for row in line_rows[1:])

            validation_rows = list(wb["Validation Results"].iter_rows(values_only=True))
            assert validation_rows[1][0] == "patient_dob"
            assert validation_rows[1][1] == "mandatory"

            evidence_rows = list(wb["Source Evidence"].iter_rows(values_only=True))
            assert any(
                row[0] == "patient_name" and row[2] == "Patient: DOE JOHN"
                for row in evidence_rows[1:]
            )

            audit_rows = list(wb["Audit Trail"].iter_rows(values_only=True))
            assert len(audit_rows) > 1


def test_excel_export_404_for_unknown_package():
    with TestClient(app) as client:
        resp = client.get("/packages/does-not-exist/export.xlsx")
        assert resp.status_code == 404
